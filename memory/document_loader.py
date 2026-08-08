"""
KinnyCode — Document Loader & Chunker

Loads documents from various formats (PDF, Markdown, plain text) and
splits them into overlapping chunks suitable for indexing in the
multi-layer memory system.

Supports:
    - PDF loading via PyPDF2 (primary) with pdfplumber fallback
    - Markdown parsing by heading-based sections
    - Plain text with encoding detection (utf-8, latin-1 fallback)
    - RecursiveCharacterTextSplitter chunking (langchain preferred, generic fallback)
    - Automatic file-type detection and ingestion orchestration

Usage:
    from core.memory.document_loader import load_document

    chunks = load_document("/path/to/doc.pdf")
    for chunk in chunks:
        print(chunk["document"][:100])
"""

from __future__ import annotations

import hashlib
import logging
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────

DEFAULT_CHUNK_SIZE = 1000
DEFAULT_CHUNK_OVERLAP = 200

# Map file extensions to document types
EXTENSION_MAP: dict[str, str] = {
    # PDF
    ".pdf": "pdf",
    # EPUB
    ".epub": "epub",
    # Word documents
    ".doc": "doc",
    ".docx": "docx",
    # OpenDocument
    ".odf": "odf",
    ".odt": "odf",
    ".ods": "odf",
    # Spreadsheet
    ".xls": "xls",
    ".xlsx": "xlsx",
    ".csv": "csv",
    # Text formats
    ".md": "md",
    ".markdown": "md",
    ".txt": "txt",
    ".rst": "txt",
    ".tex": "txt",
    ".html": "txt",
    ".htm": "txt",
    ".xml": "txt",
    ".json": "txt",
    ".yaml": "txt",
    ".yml": "txt",
    ".log": "txt",
}

# ── Generic text splitter (fallback) ──────────────────────────────────


def _generic_split(text: str, chunk_size: int, chunk_overlap: int) -> list[str]:
    """Generic character-based text splitter.

    Used as fallback when langchain_text_splitters is not available.

    Args:
        text: The text to split.
        chunk_size: Maximum chunk size in characters.
        chunk_overlap: Overlap between consecutive chunks.

    Returns:
        List of text chunks.
    """
    if not text.strip():
        return []

    chunks: list[str] = []
    start = 0

    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]

        # Try to break at a natural boundary within the chunk
        for sep in ["\n\n", "\n", ". ", "? ", "! "]:
            last = chunk.rfind(sep)
            if last > chunk_size * 0.4:  # Only use if separator is past 40% of chunk
                chunk = chunk[: last + len(sep)]
                break

        chunks.append(chunk.strip())
        start = start + chunk_size - chunk_overlap

    return [c for c in chunks if c]


# ── Standalone chunking function (compatible with memory_server._split_text) ──


def split_text(
    content: str,
    language: str = "text",
    *,
    chunk_size: int = DEFAULT_CHUNK_SIZE,
    chunk_overlap: int = DEFAULT_CHUNK_OVERLAP,
) -> list[str]:
    """Split text into overlapping chunks using RecursiveCharacterTextSplitter.

    Attempts language-specific splitting via langchain_text_splitters
    when available; falls back to generic splitting otherwise.

    This function mirrors memory_server.py's ``_split_text`` signature
    for drop-in compatibility.

    Args:
        content: The text to split.
        language: Programming language for language-aware splitting.
                  Defaults to ``"text"`` for plain text.
        chunk_size: Maximum chunk size in characters.
        chunk_overlap: Overlap between consecutive chunks.

    Returns:
        List of text chunks.
    """
    try:
        from langchain_text_splitters import RecursiveCharacterTextSplitter

        try:
            splitter = RecursiveCharacterTextSplitter.from_language(
                language=language,
                chunk_size=chunk_size,
                chunk_overlap=chunk_overlap,
            )
        except ValueError:
            splitter = RecursiveCharacterTextSplitter(
                chunk_size=chunk_size,
                chunk_overlap=chunk_overlap,
                separators=["\n\n", "\n", ". ", "? ", "! ", " ", ""],
            )
        return splitter.split_text(content)
    except ImportError:
        logger.debug("langchain_text_splitters not available; using generic splitter")
        return _generic_split(content, chunk_size, chunk_overlap)


# ═══════════════════════════════════════════════════════════════════════
#  PDF Loader
# ═══════════════════════════════════════════════════════════════════════


class PDFLoader:
    """Loads PDF files and extracts text page by page.

    Uses PyPDF2 as the primary backend; falls back to pdfplumber
    if PyPDF2 is unavailable or fails to extract text.

    Usage:
        loader = PDFLoader()
        pages = loader.load("/path/to/document.pdf")
        for page in pages:
            print(page["page"], len(page["text"]))
    """

    def load(self, file_path: str) -> list[dict[str, Any]]:
        """Load a PDF file and extract text for each page.

        Args:
            file_path: Absolute or relative path to the PDF file.

        Returns:
            List of dicts with keys ``page`` (int) and ``text`` (str).

        Raises:
            ImportError: If neither PyPDF2 nor pdfplumber is installed.
            FileNotFoundError: If the file does not exist.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"PDF file not found: {file_path}")

        # ── Try PyPDF2 first ───────────────────────────────────────
        try:
            return self._load_with_pypdf2(str(path))
        except ImportError:
            logger.debug("PyPDF2 not available; falling back to pdfplumber")
        except Exception as exc:
            logger.warning("PyPDF2 failed for %s: %s; trying pdfplumber", file_path, exc)

        # ── Fallback: pdfplumber ───────────────────────────────────
        try:
            return self._load_with_pdfplumber(str(path))
        except ImportError:
            raise ImportError(
                "No PDF library available. Install PyPDF2 or pdfplumber: "
                "pip install PyPDF2 pdfplumber"
            )
        except Exception as exc:
            raise RuntimeError(f"Failed to load PDF {file_path}: {exc}") from exc

    @staticmethod
    def _load_with_pypdf2(file_path: str) -> list[dict[str, Any]]:
        """Extract text from PDF using PyPDF2.

        Args:
            file_path: Path to the PDF file.

        Returns:
            List of dicts with ``page`` and ``text``.
        """
        from PyPDF2 import PdfReader

        reader = PdfReader(file_path)
        pages: list[dict[str, Any]] = []

        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if text:
                pages.append({"page": i + 1, "text": text.strip()})

        return pages

    @staticmethod
    def _load_with_pdfplumber(file_path: str) -> list[dict[str, Any]]:
        """Extract text from PDF using pdfplumber.

        Args:
            file_path: Path to the PDF file.

        Returns:
            List of dicts with ``page`` and ``text``.
        """
        import pdfplumber

        pages: list[dict[str, Any]] = []

        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if text:
                    pages.append({"page": i + 1, "text": text.strip()})

        return pages


# ═══════════════════════════════════════════════════════════════════════
#  EPUB Loader
# ═══════════════════════════════════════════════════════════════════════


class EPUBLoader:
    """Loads EPUB files and extracts text from chapters.

    Uses ebooklib to parse EPUB format and extract text content
    from each chapter/item in the book.

    Usage:
        loader = EPUBLoader()
        chapters = loader.load("/path/to/book.epub")
        for chapter in chapters:
            print(chapter["section"], len(chapter["text"]))
    """

    def load(self, file_path: str) -> list[dict[str, Any]]:
        """Load an EPUB file and extract text from each chapter.

        Args:
            file_path: Absolute or relative path to the EPUB file.

        Returns:
            List of dicts with keys ``section`` (str) and ``text`` (str).

        Raises:
            ImportError: If ebooklib is not installed.
            FileNotFoundError: If the file does not exist.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"EPUB file not found: {file_path}")

        try:
            import ebooklib
            from ebooklib import epub
            from bs4 import BeautifulSoup
        except ImportError:
            raise ImportError(
                "EPUB loading requires ebooklib and beautifulsoup4. "
                "Install with: pip install ebooklib beautifulsoup4"
            )

        book = epub.read_epub(str(path), options={"ignore_ncx": True})
        chapters: list[dict[str, Any]] = []

        for item in book.get_items():
            if item.get_type() == ebooklib.ITEM_DOCUMENT:
                content = item.get_content().decode("utf-8", errors="replace")
                # Parse HTML content
                soup = BeautifulSoup(content, "html.parser")
                text = soup.get_text(separator="\n", strip=True)
                
                if text.strip():
                    # Use file_name as section identifier
                    section = item.get_name() or f"Chapter {len(chapters) + 1}"
                    chapters.append({"section": section, "text": text.strip()})

        return chapters


# ═══════════════════════════════════════════════════════════════════════
#  Word Document Loader (doc/docx)
# ═══════════════════════════════════════════════════════════════════════


class WordLoader:
    """Loads Word documents (.doc, .docx) and extracts text.

    Uses python-docx for .docx files. For .doc files, attempts to
    use antiword or falls back to reading as text.

    Usage:
        loader = WordLoader()
        pages = loader.load("/path/to/document.docx")
    """

    def load(self, file_path: str) -> list[dict[str, Any]]:
        """Load a Word document and extract text.

        Args:
            file_path: Path to the .doc or .docx file.

        Returns:
            List of dicts with ``page`` and ``text``.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Word file not found: {file_path}")

        ext = path.suffix.lower()

        if ext == ".docx":
            return self._load_docx(str(path))
        elif ext == ".doc":
            return self._load_doc(str(path))
        else:
            raise ValueError(f"Unsupported Word extension: {ext}")

    @staticmethod
    def _load_docx(file_path: str) -> list[dict[str, Any]]:
        """Extract text from .docx using python-docx."""
        try:
            from docx import Document
        except ImportError:
            raise ImportError(
                "DOCX loading requires python-docx. "
                "Install with: pip install python-docx"
            )

        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        full_text = "\n".join(paragraphs)
        
        if not full_text.strip():
            return []
        
        return [{"page": 1, "text": full_text.strip()}]

    @staticmethod
    def _load_doc(file_path: str) -> list[dict[str, Any]]:
        """Extract text from .doc (legacy format)."""
        try:
            # Try using antiword if available
            import subprocess
            result = subprocess.run(
                ["antiword", file_path],
                capture_output=True,
                text=True,
                timeout=30
            )
            if result.returncode == 0 and result.stdout.strip():
                return [{"page": 1, "text": result.stdout.strip()}]
        except (FileNotFoundError, subprocess.TimeoutExpired):
            pass

        # Fallback: try reading as plain text with encoding detection
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                text = f.read()
            if text.strip():
                return [{"page": 1, "text": text.strip()}]
        except Exception:
            pass

        raise RuntimeError(
            f"Cannot read .doc file: {file_path}. "
            "Install antiword or convert to .docx first."
        )


# ═══════════════════════════════════════════════════════════════════════
#  OpenDocument Loader (odf/odt/ods)
# ═══════════════════════════════════════════════════════════════════════


class OpenDocumentLoader:
    """Loads OpenDocument files (.odf, .odt, .ods) and extracts text.

    Uses odfpy for parsing OpenDocument format.

    Usage:
        loader = OpenDocumentLoader()
        pages = loader.load("/path/to/document.odt")
    """

    def load(self, file_path: str) -> list[dict[str, Any]]:
        """Load an OpenDocument file and extract text.

        Args:
            file_path: Path to the .odf, .odt, or .ods file.

        Returns:
            List of dicts with ``page`` and ``text``.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"OpenDocument file not found: {file_path}")

        try:
            from odf import text, teletype
            from odf.opendocument import load
        except ImportError:
            raise ImportError(
                "OpenDocument loading requires odfpy. "
                "Install with: pip install odfpy"
            )

        doc = load(str(path))
        paragraphs = []
        
        for paragraph in doc.getElementsByType(text.P):
            txt = teletype.extractText(paragraph)
            if txt.strip():
                paragraphs.append(txt.strip())

        full_text = "\n".join(paragraphs)
        
        if not full_text.strip():
            return []
        
        return [{"page": 1, "text": full_text.strip()}]


# ═══════════════════════════════════════════════════════════════════════
#  Spreadsheet Loader (xls/xlsx/csv)
# ═══════════════════════════════════════════════════════════════════════


class SpreadsheetLoader:
    """Loads spreadsheet files (.xls, .xlsx, .csv) and extracts text.

    Uses openpyxl for .xlsx, xlrd for .xls, and csv module for .csv.

    Usage:
        loader = SpreadsheetLoader()
        pages = loader.load("/path/to/data.xlsx")
    """

    def load(self, file_path: str) -> list[dict[str, Any]]:
        """Load a spreadsheet file and extract text.

        Args:
            file_path: Path to the .xls, .xlsx, or .csv file.

        Returns:
            List of dicts with ``page`` (sheet name) and ``text``.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Spreadsheet file not found: {file_path}")

        ext = path.suffix.lower()

        if ext == ".xlsx":
            return self._load_xlsx(str(path))
        elif ext == ".xls":
            return self._load_xls(str(path))
        elif ext == ".csv":
            return self._load_csv(str(path))
        else:
            raise ValueError(f"Unsupported spreadsheet extension: {ext}")

    @staticmethod
    def _load_xlsx(file_path: str) -> list[dict[str, Any]]:
        """Extract text from .xlsx using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "XLSX loading requires openpyxl. "
                "Install with: pip install openpyxl"
            )

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        pages = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    rows.append(row_text)
            
            if rows:
                text = "\n".join(rows)
                pages.append({"page": sheet_name, "text": text.strip()})

        wb.close()
        return pages

    @staticmethod
    def _load_xls(file_path: str) -> list[dict[str, Any]]:
        """Extract text from .xls using xlrd."""
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "XLS loading requires xlrd. "
                "Install with: pip install xlrd"
            )

        wb = xlrd.open_workbook(file_path)
        pages = []

        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            rows = []
            for row_idx in range(ws.nrows):
                row = [str(ws.cell_value(row_idx, col_idx)) for col_idx in range(ws.ncols)]
                row_text = " | ".join(row)
                if row_text.strip():
                    rows.append(row_text)
            
            if rows:
                text = "\n".join(rows)
                pages.append({"page": ws.name, "text": text.strip()})

        return pages

    @staticmethod
    def _load_csv(file_path: str) -> list[dict[str, Any]]:
        """Extract text from .csv using csv module."""
        import csv
        
        pages = []
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                rows = [", ".join(row) for row in reader if any(cell.strip() for cell in row)]
            
            if rows:
                text = "\n".join(rows)
                pages.append({"page": 1, "text": text.strip()})
        except Exception as e:
            raise RuntimeError(f"Failed to read CSV: {e}")

        return pages


# ═══════════════════════════════════════════════════════════════════════
#  Markdown Loader
# ═══════════════════════════════════════════════════════════════════════

_HEADING_RE = re.compile(r"^#{1,3}\s+")


class MarkdownLoader:
    """Loads Markdown files and parses headings as section boundaries.

    Recognizes headings at levels 1-3 (``#``, ``##``, ``###``) as
    section delimiters. Text before the first heading is assigned
    an empty section name.

    Usage:
        loader = MarkdownLoader()
        sections = loader.load("/path/to/readme.md")
        for section in sections:
            print(section["section"], len(section["text"]))
    """

    def load(self, file_path: str) -> list[dict[str, Any]]:
        """Load and parse a Markdown file into heading-based sections.

        Args:
            file_path: Absolute or relative path to the Markdown file.

        Returns:
            List of dicts with keys ``section`` (str) and ``text`` (str).

        Raises:
            FileNotFoundError: If the file does not exist.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Markdown file not found: {file_path}")

        with open(str(path), "r", encoding="utf-8", errors="replace") as f:
            content = f.read()

        return self._parse_sections(content)

    @staticmethod
    def load_from_text(content: str) -> list[dict[str, Any]]:
        """Parse markdown content into heading-based sections.

        Args:
            content: Raw markdown string.

        Returns:
            List of dicts with keys ``section`` and ``text``.
        """
        return MarkdownLoader._parse_sections(content)

    @staticmethod
    def _parse_sections(content: str) -> list[dict[str, Any]]:
        """Parse markdown text into heading-based sections.

        Args:
            content: Raw markdown content.

        Returns:
            List of sections with ``section`` and ``text`` keys.
        """
        lines = content.split("\n")
        sections: list[dict[str, Any]] = []
        current_heading = ""
        current_lines: list[str] = []

        for line in lines:
            stripped = line.lstrip()

            if _HEADING_RE.match(stripped):
                # Save previous section
                text = "\n".join(current_lines).strip()
                if text or current_heading:
                    sections.append({
                        "section": current_heading,
                        "text": text,
                    })

                # Start new section
                current_heading = stripped.lstrip("#").strip()
                current_lines = []
            else:
                current_lines.append(line)

        # Save last section
        text = "\n".join(current_lines).strip()
        if text or current_heading:
            sections.append({
                "section": current_heading,
                "text": text,
            })

        return sections


# ═══════════════════════════════════════════════════════════════════════
#  Text Loader
# ═══════════════════════════════════════════════════════════════════════


class TextLoader:
    """Loads plain text files with automatic encoding detection.

    Attempts UTF-8 first; falls back to latin-1 on decode errors.
    This handles most real-world encoding issues without external
    dependencies.

    Usage:
        loader = TextLoader()
        result = loader.load("/path/to/file.txt")
        print(result[0]["text"][:100])
    """

    def load(self, file_path: str) -> list[dict[str, Any]]:
        """Load a plain text file with encoding detection.

        Args:
            file_path: Absolute or relative path to the text file.

        Returns:
            List with a single dict containing key ``text`` (str).

        Raises:
            FileNotFoundError: If the file does not exist.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Text file not found: {file_path}")

        text = self._read_with_encoding_fallback(str(path))
        return [{"text": text}]

    @staticmethod
    def _read_with_encoding_fallback(file_path: str) -> str:
        """Read file with UTF-8 primary and latin-1 fallback.

        Reads raw bytes and attempts UTF-8 decode (strict). If that
        fails due to invalid byte sequences, falls back to latin-1
        with replacement characters for truly unrepresentable bytes.

        Args:
            file_path: Path to the file.

        Returns:
            Decoded string content.
        """
        with open(file_path, "rb") as f:
            raw_bytes = f.read()

        try:
            return raw_bytes.decode("utf-8")
        except UnicodeDecodeError:
            logger.debug(
                "UTF-8 decode failed for %s; falling back to latin-1",
                file_path,
            )
            return raw_bytes.decode("latin-1", errors="replace")


# ═══════════════════════════════════════════════════════════════════════
#  Document Chunker
# ═══════════════════════════════════════════════════════════════════════


class DocumentChunker:
    """Splits document text into overlapping chunks for indexing.

    Uses langchain's ``RecursiveCharacterTextSplitter`` when available,
    with a generic character-based splitter as fallback. Each chunk
    is enriched with metadata (source file, document type, page number,
    content hash) suitable for the multi-layer memory system.

    Usage:
        chunker = DocumentChunker(chunk_size=1000, chunk_overlap=200)
        chunks = chunker.chunk(
            text="Long document text...",
            source_file="/path/to/doc.pdf",
            doc_type="pdf",
            page_number=1,
        )
        for chunk in chunks:
            print(chunk["id"], len(chunk["document"]))
    """

    def __init__(
        self,
        chunk_size: int = DEFAULT_CHUNK_SIZE,
        chunk_overlap: int = DEFAULT_CHUNK_OVERLAP,
    ) -> None:
        """Initialize the document chunker.

        Args:
            chunk_size: Maximum chunk size in characters (default 1000).
            chunk_overlap: Overlap between consecutive chunks (default 200).
        """
        self._chunk_size = chunk_size
        self._chunk_overlap = chunk_overlap

    def chunk(
        self,
        text: str,
        source_file: str,
        doc_type: str,
        page_number: int | None = None,
    ) -> list[dict[str, Any]]:
        """Split text into overlapping chunks with full metadata.

        Args:
            text: The raw document text to split.
            source_file: Original file path (used for chunk IDs).
            doc_type: Document type (``"pdf"``, ``"md"``, ``"txt"``).
            page_number: Optional page number for PDF documents.

        Returns:
            List of chunk dicts, each with keys:
                - ``id``: Unique chunk identifier.
                - ``document``: Chunk text content.
                - ``source_file``: Original file path.
                - ``doc_type``: Document type.
                - ``chunk_index``: Zero-based chunk index.
                - ``page_number``: Page number or None.
                - ``content_hash``: SHA256 hex digest of chunk text.
        """
        if not text or not text.strip():
            return []

        # Split text into chunks
        text_chunks = split_text(
            text,
            language="text",
            chunk_size=self._chunk_size,
            chunk_overlap=self._chunk_overlap,
        )

        # Enrich with metadata
        result: list[dict[str, Any]] = []
        for i, chunk_text in enumerate(text_chunks):
            chunk_id = f"{source_file}_chunk_{i}"
            chunk_hash = hashlib.sha256(chunk_text.encode("utf-8")).hexdigest()

            result.append({
                "id": chunk_id,
                "document": chunk_text,
                "source_file": source_file,
                "doc_type": doc_type,
                "chunk_index": i,
                "page_number": page_number,
                "content_hash": chunk_hash,
            })

        return result

    def chunk_texts(
        self,
        texts: list[str],
        source_file: str,
        doc_type: str,
    ) -> list[dict[str, Any]]:
        """Chunk multiple text segments from a single source document.

        Useful for PDF pages or markdown sections that should be
        chunked separately but share the same source file metadata.

        Args:
            texts: List of text segments to chunk.
            source_file: Original file path.
            doc_type: Document type.

        Returns:
            Combined list of chunk dicts from all text segments.
        """
        all_chunks: list[dict[str, Any]] = []
        for text in texts:
            if text and text.strip():
                chunks = self.chunk(
                    text,
                    source_file=source_file,
                    doc_type=doc_type,
                )
                all_chunks.extend(chunks)
        return all_chunks


# ═══════════════════════════════════════════════════════════════════════
#  Document Ingestor (Orchestrator)
# ═══════════════════════════════════════════════════════════════════════


class DocumentIngestor:
    """Orchestrates document loading and chunking.

    Auto-detects file type by extension, selects the appropriate loader,
    loads the content, and chunks it using ``DocumentChunker``.

    Supports two ingestion modes:
        1. File-based: reads from disk with ``ingest(file_path)``.
        2. Content-based: accepts pre-loaded content with
           ``ingest_content(file_path, content, doc_type)``.

    Usage:
        ingestor = DocumentIngestor()
        chunks = ingestor.ingest("/path/to/document.pdf")
        print(f"Ingested {len(chunks)} chunks")

        # Or with pre-loaded content:
        chunks = ingestor.ingest_content(
            "/docs/notes.md", markdown_content, "md"
        )
    """

    def __init__(
        self,
        chunk_size: int = DEFAULT_CHUNK_SIZE,
        chunk_overlap: int = DEFAULT_CHUNK_OVERLAP,
    ) -> None:
        """Initialize the document ingestor.

        Args:
            chunk_size: Maximum chunk size for the internal DocumentChunker.
            chunk_overlap: Chunk overlap for the internal DocumentChunker.
        """
        self._chunker = DocumentChunker(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )
        self._pdf_loader = PDFLoader()
        self._md_loader = MarkdownLoader()
        self._text_loader = TextLoader()
        self._epub_loader = EPUBLoader()
        self._word_loader = WordLoader()
        self._odf_loader = OpenDocumentLoader()
        self._spreadsheet_loader = SpreadsheetLoader()

    def ingest(self, file_path: str) -> list[dict[str, Any]]:
        """Load and chunk a document file.

        Detects the document type from the file extension and
        uses the appropriate loader followed by chunking.

        Args:
            file_path: Path to the document file.

        Returns:
            List of chunk dicts with full metadata.

        Raises:
            ValueError: If the file extension is not supported.
            FileNotFoundError: If the file does not exist.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        doc_type = self._detect_type(path)
        ext = path.suffix.lower()

        if doc_type == "pdf":
            return self._ingest_pdf(str(path))

        elif doc_type == "md":
            return self._ingest_markdown(str(path))

        elif doc_type == "txt":
            return self._ingest_text(str(path))

        elif doc_type == "epub":
            return self._ingest_epub(str(path))

        elif doc_type in ("doc", "docx"):
            return self._ingest_word(str(path))

        elif doc_type == "odf":
            return self._ingest_odf(str(path))

        elif doc_type in ("xls", "xlsx", "csv"):
            return self._ingest_spreadsheet(str(path))

        else:
            raise ValueError(
                f"Unsupported document type for extension '{ext}'. "
                f"Supported: {', '.join(sorted(EXTENSION_MAP.keys()))}"
            )

    def ingest_content(
        self,
        file_path: str,
        content: str,
        doc_type: str,
    ) -> list[dict[str, Any]]:
        """Chunk pre-loaded document content.

        Useful when content is obtained from a source other than disk
        (e.g., API response, database, memory buffer).

        Args:
            file_path: Virtual file path used for chunk IDs and metadata.
            content: The raw document text.
            doc_type: Document type (``"pdf"``, ``"md"``, ``"txt"``).

        Returns:
            List of chunk dicts with full metadata.

        Raises:
            ValueError: If ``doc_type`` is unrecognized.
        """
        if not content or not content.strip():
            return []

        doc_type = doc_type.lower().strip()

        if doc_type == "md":
            return self._ingest_markdown_content(file_path, content)

        # Treat all other types as plain text
        return self._chunker.chunk(
            text=content,
            source_file=file_path,
            doc_type=doc_type,
        )

    # ── Private ingestion helpers ──────────────────────────────────

    def _ingest_pdf(self, file_path: str) -> list[dict[str, Any]]:
        """Load PDF text page by page and chunk each page.

        Args:
            file_path: Path to the PDF file.

        Returns:
            List of chunk dicts with page_number metadata.
        """
        pages = self._pdf_loader.load(file_path)
        all_chunks: list[dict[str, Any]] = []

        for page in pages:
            page_num = page["page"]
            text = page["text"]

            chunks = self._chunker.chunk(
                text=text,
                source_file=file_path,
                doc_type="pdf",
                page_number=page_num,
            )
            all_chunks.extend(chunks)

        return all_chunks

    def _ingest_markdown(self, file_path: str) -> list[dict[str, Any]]:
        """Load Markdown file, parse sections, and chunk.

        Args:
            file_path: Path to the Markdown file.

        Returns:
            List of chunk dicts.
        """
        sections = self._md_loader.load(file_path)
        all_chunks: list[dict[str, Any]] = []

        for section in sections:
            text = section["text"]
            if text and text.strip():
                chunks = self._chunker.chunk(
                    text=text,
                    source_file=file_path,
                    doc_type="md",
                )
                all_chunks.extend(chunks)

        return all_chunks

    def _ingest_markdown_content(
        self, file_path: str, content: str
    ) -> list[dict[str, Any]]:
        """Parse markdown content and chunk each section.

        Args:
            file_path: Virtual file path for metadata.
            content: Raw markdown string.

        Returns:
            List of chunk dicts.
        """
        sections = MarkdownLoader.load_from_text(content)
        all_chunks: list[dict[str, Any]] = []

        for section in sections:
            text = section["text"]
            if text and text.strip():
                chunks = self._chunker.chunk(
                    text=text,
                    source_file=file_path,
                    doc_type="md",
                )
                all_chunks.extend(chunks)

        return all_chunks

    def _ingest_text(self, file_path: str) -> list[dict[str, Any]]:
        """Load plain text file and chunk as a single document.

        Args:
            file_path: Path to the text file.

        Returns:
            List of chunk dicts.
        """
        result = self._text_loader.load(file_path)
        if not result:
            return []

        text = result[0]["text"]
        if not text or not text.strip():
            return []

        return self._chunker.chunk(
            text=text,
            source_file=file_path,
            doc_type="txt",
        )

    def _ingest_epub(self, file_path: str) -> list[dict[str, Any]]:
        """Load EPUB file and chunk each chapter.

        Args:
            file_path: Path to the EPUB file.

        Returns:
            List of chunk dicts.
        """
        chapters = self._epub_loader.load(file_path)
        all_chunks = []

        for chapter in chapters:
            text = chapter.get("text", "")
            if text.strip():
                chunks = self._chunker.chunk(
                    text=text,
                    source_file=file_path,
                    doc_type="epub",
                )
                all_chunks.extend(chunks)

        return all_chunks

    def _ingest_word(self, file_path: str) -> list[dict[str, Any]]:
        """Load Word document and chunk as a single document.

        Args:
            file_path: Path to the .doc or .docx file.

        Returns:
            List of chunk dicts.
        """
        result = self._word_loader.load(file_path)
        if not result:
            return []

        text = result[0]["text"]
        if not text or not text.strip():
            return []

        ext = Path(file_path).suffix.lower()
        doc_type = "docx" if ext == ".docx" else "doc"

        return self._chunker.chunk(
            text=text,
            source_file=file_path,
            doc_type=doc_type,
        )

    def _ingest_odf(self, file_path: str) -> list[dict[str, Any]]:
        """Load OpenDocument file and chunk as a single document.

        Args:
            file_path: Path to the .odf/.odt/.ods file.

        Returns:
            List of chunk dicts.
        """
        result = self._odf_loader.load(file_path)
        if not result:
            return []

        text = result[0]["text"]
        if not text or not text.strip():
            return []

        return self._chunker.chunk(
            text=text,
            source_file=file_path,
            doc_type="odf",
        )

    def _ingest_spreadsheet(self, file_path: str) -> list[dict[str, Any]]:
        """Load spreadsheet file and chunk each sheet.

        Args:
            file_path: Path to the .xls/.xlsx/.csv file.

        Returns:
            List of chunk dicts.
        """
        sheets = self._spreadsheet_loader.load(file_path)
        all_chunks = []

        for sheet in sheets:
            text = sheet.get("text", "")
            if text.strip():
                ext = Path(file_path).suffix.lower()
                doc_type = "csv" if ext == ".csv" else ("xlsx" if ext == ".xlsx" else "xls")
                chunks = self._chunker.chunk(
                    text=text,
                    source_file=file_path,
                    doc_type=doc_type,
                )
                all_chunks.extend(chunks)

        return all_chunks

    @staticmethod
    def _detect_type(path: Path) -> str:
        """Detect the document type from the file extension.

        Args:
            path: Path object for the file.

        Returns:
            Document type string (``"pdf"``, ``"md"``, ``"txt"``),
            or ``"unknown"`` if unrecognized.
        """
        ext = path.suffix.lower()
        return EXTENSION_MAP.get(ext, "unknown")


# ═══════════════════════════════════════════════════════════════════════
#  Module-level convenience function
# ═══════════════════════════════════════════════════════════════════════


def load_document(
    file_path: str,
    chunk_size: int = DEFAULT_CHUNK_SIZE,
    chunk_overlap: int = DEFAULT_CHUNK_OVERLAP,
) -> list[dict[str, Any]]:
    """Load and chunk a document file.

    Convenience function that creates a ``DocumentIngestor``
    and calls ``ingest(file_path)``. Suitable for quick one-off
    document processing.

    Args:
        file_path: Path to the document file (``.pdf``, ``.md``, ``.txt``, etc.).
        chunk_size: Maximum chunk size in characters (default 1000).
        chunk_overlap: Overlap between consecutive chunks (default 200).

    Returns:
        List of chunk dicts, each containing:
            - ``id``: Unique chunk identifier (``"{path}_chunk_{i}"``).
            - ``document``: Chunk text.
            - ``source_file``: Original file path.
            - ``doc_type``: Document type (``"pdf"``, ``"md"``, ``"txt"``).
            - ``chunk_index``: Zero-based index within the document.
            - ``page_number``: Page number for PDFs, ``None`` otherwise.
            - ``content_hash``: SHA256 hex digest of chunk text.

    Example:
        >>> from core.memory.document_loader import load_document
        >>> chunks = load_document("/docs/readme.md")
        >>> for c in chunks:
        ...     print(c["id"], len(c["document"]))
    """
    ingestor = DocumentIngestor(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
    )
    return ingestor.ingest(file_path)
